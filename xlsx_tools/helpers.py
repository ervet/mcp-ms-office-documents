import re
import logging
from dataclasses import dataclass

from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

# ── Layout Constants (shared with base_xlsx_tool) ─────────────────────────────
TABLE_BOTTOM_SPACING = 2
MIN_COLUMN_WIDTH = 12
MAX_COLUMN_WIDTH = 25
COLUMN_WIDTH_PADDING = 2


def parse_table(lines: list[str], start_idx: int) -> tuple[list[list[str]] | None, int]:
    """Parse markdown table and return (table_data, next_index)."""
    table_lines: list[str] = []
    i = start_idx

    # Find all consecutive table lines
    while i < len(lines):
        line = lines[i].strip()
        if line.startswith('|') and line.endswith('|'):
            table_lines.append(line)
            i += 1
        else:
            break

    if len(table_lines) < 2:  # Need at least header and separator
        return None, start_idx + 1

    # Parse table data skipping separator row
    table_data: list[list[str]] = []
    for line in table_lines:
        if '---' in line or ':-:' in line or ':--' in line or '--:' in line:
            continue
        cells = [cell.strip() for cell in line.split('|')[1:-1]]
        table_data.append(cells)

    return table_data, i


# ── Cell Resolution ───────────────────────────────────────────────────────────

@dataclass
class CellResult:
    """Resolved cell metadata — all information needed to write a cell to Excel."""
    value: str | int | float  # The cleaned value to write
    is_formula: bool = False
    is_percent: bool = False
    bold: bool = False
    italic: bool = False
    monospace: bool = False

    @property
    def formatting_info(self) -> dict[str, bool]:
        """Legacy-compatible formatting dict for apply_cell_formatting()."""
        return {'bold': self.bold, 'italic': self.italic, 'monospace': self.monospace}


def _detect_implicit_formula(value: str) -> str | None:
    """Detect common formula patterns written without '=' prefix.

    Returns the formula string (with '=' prepended) or None if not a formula pattern.
    """
    if re.match(r'^(SUM|sum)\([A-Z]+\d+:[A-Z]+\d+\)$', value):
        return f"={value.upper()}"
    if re.match(r'^(AVG|avg|AVERAGE|average)\([A-Z]+\d+:[A-Z]+\d+\)$', value):
        return f"=AVERAGE({value.split('(')[1]}"
    if re.match(r'^[A-Z]+\d+[+\-*/][A-Z]+\d+$', value):
        return f"={value}"
    return None


def resolve_cell(raw_text: str) -> CellResult:
    """Parse a raw markdown cell string into a fully resolved CellResult.

    Combines formatting detection, formula detection, and type conversion
    in a single pass — the unified replacement for the former three-function pipeline
    of parse_cell_formatting → detect_formula_pattern → format_cell_value.
    """
    # Step 1: Strip markdown formatting markers
    clean_text = raw_text.strip()
    bold = False
    italic = False
    monospace = False

    if clean_text.startswith('**') and clean_text.endswith('**') and len(clean_text) > 4:
        clean_text = clean_text[2:-2]
        bold = True
    elif clean_text.startswith('*') and clean_text.endswith('*') and len(clean_text) > 2:
        clean_text = clean_text[1:-1]
        italic = True
    elif clean_text.startswith('`') and clean_text.endswith('`') and len(clean_text) > 2:
        clean_text = clean_text[1:-1]
        monospace = True

    # Step 2: Check if it's an explicit formula (= prefix)
    if clean_text.startswith('='):
        return CellResult(
            value=clean_text, is_formula=True,
            bold=bold, italic=italic, monospace=monospace,
        )

    # Step 3: Detect implicit formula patterns (without = prefix)
    implicit_formula = _detect_implicit_formula(clean_text)
    if implicit_formula:
        return CellResult(
            value=implicit_formula, is_formula=True,
            bold=bold, italic=italic, monospace=monospace,
        )

    # Step 4: Detect percent and convert to number
    is_percent = clean_text.endswith('%')
    if is_percent:
        try:
            numeric_val = float(clean_text[:-1]) / 100
            return CellResult(
                value=numeric_val, is_percent=True,
                bold=bold, italic=italic, monospace=monospace,
            )
        except ValueError:
            pass  # Not a valid percent number — fall through to text

    # Step 5: Try numeric conversion
    try:
        numeric_val = float(clean_text)
        return CellResult(
            value=numeric_val,
            bold=bold, italic=italic, monospace=monospace,
        )
    except ValueError:
        pass

    # Step 6: Plain text
    return CellResult(
        value=clean_text,
        bold=bold, italic=italic, monospace=monospace,
    )


def apply_cell_formatting(cell, formatting_info: dict[str, bool]) -> None:
    """Apply formatting information to an Excel cell."""
    current_font = cell.font
    if formatting_info['bold']:
        cell.font = Font(bold=True, color=current_font.color, size=current_font.size)
    elif formatting_info['italic']:
        cell.font = Font(italic=True, color=current_font.color, size=current_font.size)
    elif formatting_info['monospace']:
        cell.font = Font(name='Courier New', color=current_font.color, size=current_font.size)


# ── Formula Reference Resolution ─────────────────────────────────────────────

def _quote_sheet_name(name: str) -> str:
    """Return the sheet name quoted for Excel if it contains spaces or special chars."""
    if re.search(r"[^A-Za-z0-9_]", name):
        return f"'{name}'"
    return name


def _resolve_row(positions: dict[str, int], table_num: int, offset: int, fallback_row: int) -> int:
    """Resolve a table-relative row reference to an absolute Excel row number.

    Args:
        positions: Table positions dict ({"T1": start_row, ...}) for the target sheet.
        table_num: Table number (1-based).
        offset: Row offset within the table (0 = first data row).
        fallback_row: Row to use if the table isn't found in positions.

    Returns:
        The absolute Excel row number.
    """
    key = f"T{table_num}"
    base = positions.get(key)
    if base is not None:
        return base + 1 + offset  # +1 to skip header row
    return fallback_row + offset


def _make_cell_ref(column: str, row: int, sheet: str | None = None) -> str:
    """Build a cell reference string, optionally with a quoted sheet prefix."""
    if sheet:
        return f"{_quote_sheet_name(sheet)}!{column}{row}"
    return f"{column}{row}"


def adjust_formula_references(
    formula: str,
    current_excel_row: int,
    table_positions: dict[str, int] | None = None,
    all_sheet_table_positions: dict[str, dict[str, int]] | None = None,
) -> str:
    """Convert row-relative references [offset] and table references T1.B[1] to actual Excel row numbers.

    Also resolves cross-sheet references like ``SheetName!T1.B[0]`` → ``'SheetName'!B2``.
    """
    if not formula.startswith('='):
        return formula

    if table_positions is None:
        table_positions = {}
    if all_sheet_table_positions is None:
        all_sheet_table_positions = {}

    logger.debug("Resolving formula: %s (current_row=%d)", formula, current_excel_row)

    try:
        # ── Cross-sheet references (must be resolved BEFORE local patterns) ──

        # Cross-sheet function: SheetName!T1.SUM(B[0]:E[0])
        cs_func_pattern = r"([\w\s.]+)!T(\d+)\.(SUM|AVERAGE|MAX|MIN)\(([A-Z]+)\[([+-]?\d+)\]:([A-Z]+)\[([+-]?\d+)\]\)"

        def _replace_cs_func(match):
            sheet = match.group(1).strip()
            table_num = int(match.group(2))
            func_name = match.group(3)
            start_col = match.group(4)
            start_offset = int(match.group(5))
            end_col = match.group(6)
            end_offset = int(match.group(7))
            pos = all_sheet_table_positions.get(sheet, {})
            sr = _resolve_row(pos, table_num, start_offset, current_excel_row)
            er = _resolve_row(pos, table_num, end_offset, current_excel_row)
            qs = _quote_sheet_name(sheet)
            result = f"{func_name}({qs}!{start_col}{sr}:{qs}!{end_col}{er})"
            logger.debug("  Cross-sheet func: %s → %s", match.group(0), result)
            return result

        formula = re.sub(cs_func_pattern, _replace_cs_func, formula)

        # Cross-sheet range: SheetName!T1.B[0]:T1.E[0]
        cs_range_pattern = r"([\w\s.]+)!T(\d+)\.([A-Z]+)\[([+-]?\d+)\]:T(\d+)\.([A-Z]+)\[([+-]?\d+)\]"

        def _replace_cs_range(match):
            sheet = match.group(1).strip()
            st_num = int(match.group(2))
            start_col = match.group(3)
            start_offset = int(match.group(4))
            et_num = int(match.group(5))
            end_col = match.group(6)
            end_offset = int(match.group(7))
            pos = all_sheet_table_positions.get(sheet, {})
            sr = _resolve_row(pos, st_num, start_offset, current_excel_row)
            er = _resolve_row(pos, et_num, end_offset, current_excel_row)
            qs = _quote_sheet_name(sheet)
            result = f"{qs}!{start_col}{sr}:{end_col}{er}"
            logger.debug("  Cross-sheet range: %s → %s", match.group(0), result)
            return result

        formula = re.sub(cs_range_pattern, _replace_cs_range, formula)

        # Cross-sheet single cell: SheetName!T1.B[0]
        cs_cell_pattern = r"([\w\s.]+)!T(\d+)\.([A-Z]+)\[([+-]?\d+)\]"

        def _replace_cs_cell(match):
            sheet = match.group(1).strip()
            table_num = int(match.group(2))
            column = match.group(3)
            offset = int(match.group(4))
            pos = all_sheet_table_positions.get(sheet, {})
            actual_row = _resolve_row(pos, table_num, offset, current_excel_row)
            result = _make_cell_ref(column, actual_row, sheet)
            logger.debug("  Cross-sheet cell: %s → %s", match.group(0), result)
            return result

        formula = re.sub(cs_cell_pattern, _replace_cs_cell, formula)

        # ── Local (same-sheet) references ──
        # NOTE: Range and function patterns must be processed BEFORE single-cell
        # to prevent the single-cell regex from consuming parts of range expressions.

        # Table range references e.g. T1.B[0]:T1.E[0]
        table_range_pattern = r'T(\d+)\.([A-Z]+)\[([+-]?\d+)\]:T(\d+)\.([A-Z]+)\[([+-]?\d+)\]'

        def replace_table_range(match):
            start_table_num = int(match.group(1))
            start_col = match.group(2)
            start_offset = int(match.group(3))
            end_table_num = int(match.group(4))
            end_col = match.group(5)
            end_offset = int(match.group(6))
            start_row = _resolve_row(table_positions, start_table_num, start_offset, current_excel_row)
            end_row = _resolve_row(table_positions, end_table_num, end_offset, current_excel_row)
            return f"{start_col}{start_row}:{end_col}{end_row}"

        adjusted = re.sub(table_range_pattern, replace_table_range, formula)

        # Simplified function over table range e.g. T1.SUM(B[0]:E[0])
        table_func_pattern = r'T(\d+)\.(SUM|AVERAGE|MAX|MIN)\(([A-Z]+)\[([+-]?\d+)\]:([A-Z]+)\[([+-]?\d+)\]\)'

        def replace_table_function(match):
            table_num = int(match.group(1))
            func_name = match.group(2)
            start_col = match.group(3)
            start_offset = int(match.group(4))
            end_col = match.group(5)
            end_offset = int(match.group(6))
            start_row = _resolve_row(table_positions, table_num, start_offset, current_excel_row)
            end_row = _resolve_row(table_positions, table_num, end_offset, current_excel_row)
            return f"{func_name}({start_col}{start_row}:{end_col}{end_row})"

        adjusted = re.sub(table_func_pattern, replace_table_function, adjusted)

        # Table cell references e.g. T1.B[1] (AFTER range patterns)
        table_pattern = r'T(\d+)\.([A-Z]+)\[([+-]?\d+)\]'

        def replace_table_reference(match):
            table_num = int(match.group(1))
            column = match.group(2)
            offset = int(match.group(3))
            actual_row = _resolve_row(table_positions, table_num, offset, current_excel_row)
            result = f"{column}{actual_row}"
            logger.debug("  Local table ref: %s → %s", match.group(0), result)
            return result

        adjusted = re.sub(table_pattern, replace_table_reference, adjusted)

        # Determine current table start for relative references
        current_table_start = None
        for table_key, table_start_row in table_positions.items():
            if table_start_row <= current_excel_row:
                current_table_start = table_start_row

        # Row-relative range e.g. B[0]:E[0] (BEFORE single-cell relative)
        range_pattern = r'([A-Z]+)\[([+-]?\d+)\]:([A-Z]+)\[([+-]?\d+)\]'

        def replace_range(match):
            start_col = match.group(1)
            start_offset = int(match.group(2))
            end_col = match.group(3)
            end_offset = int(match.group(4))
            if current_table_start is not None:
                start_row = current_table_start + 1 + start_offset
                end_row = current_table_start + 1 + end_offset
            else:
                start_row = current_excel_row + start_offset
                end_row = current_excel_row + end_offset
            return f"{start_col}{start_row}:{end_col}{end_row}"

        adjusted = re.sub(range_pattern, replace_range, adjusted)

        # Handle row-relative references e.g. B[0] (AFTER range pattern)
        rel_pattern = r'([A-Z]+)\[([+-]?\d+)\]'

        def replace_rel(match):
            column = match.group(1)
            offset = int(match.group(2))
            if current_table_start is not None:
                actual_row = current_table_start + 1 + offset
            else:
                actual_row = current_excel_row + offset
            result = f"{column}{actual_row}"
            logger.debug("  Relative ref: %s → %s", match.group(0), result)
            return result

        adjusted = re.sub(rel_pattern, replace_rel, adjusted)

        logger.debug("  Resolved formula: %s → %s", formula, adjusted)
        return adjusted

    except Exception as e:
        logger.warning("Failed to adjust formula references for '%s': %s", formula, e)
        return formula


# ── Table Rendering ───────────────────────────────────────────────────────────

def add_table_to_sheet(
    table_data: list[list[str]],
    worksheet,
    start_row: int,
    table_positions: dict[str, int] | None = None,
    all_sheet_table_positions: dict[str, dict[str, int]] | None = None,
) -> int:
    """Add table data to Excel worksheet with proper formatting and formula support."""
    if not table_data:
        return start_row

    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    formula_fill = PatternFill(start_color="E7F3FF", end_color="E7F3FF", fill_type="solid")
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # Fill cells
    for row_idx, row_data in enumerate(table_data):
        current_excel_row = start_row + row_idx
        for col_idx, cell_text in enumerate(row_data):
            try:
                cell = worksheet.cell(row=current_excel_row, column=col_idx + 1)
                resolved = resolve_cell(cell_text)

                if resolved.is_formula:
                    adjusted_formula = adjust_formula_references(
                        resolved.value, current_excel_row, table_positions, all_sheet_table_positions
                    )
                    cell.value = adjusted_formula
                    cell.fill = formula_fill
                else:
                    cell.value = resolved.value

                # Apply inline formatting (bold/italic/monospace)
                apply_cell_formatting(cell, resolved.formatting_info)
                cell.border = border

                # Alignment
                if row_idx == 0:
                    cell.alignment = Alignment(horizontal='center')
                elif isinstance(cell.value, (int, float)) or (isinstance(cell.value, str) and cell.value.startswith('=')):
                    cell.alignment = Alignment(horizontal='right')
                else:
                    cell.alignment = Alignment(horizontal='left')

                # Header row styling (overrides inline formatting)
                if row_idx == 0:
                    cell.font = header_font
                    cell.fill = header_fill
                elif isinstance(cell.value, (int, float)) and cell.value >= 1000:
                    cell.number_format = '#,##0'

                # Apply percentage number format
                if resolved.is_percent and isinstance(cell.value, (int, float)):
                    cell.number_format = '0%'
            except Exception as e:
                logger.warning("Error processing cell [row=%d, col=%d]: %s", current_excel_row, col_idx + 1, e)

    # Column widths
    for col_idx in range(len(table_data[0]) if table_data else 0):
        column_letter = get_column_letter(col_idx + 1)
        max_length = 0
        for row in table_data:
            if col_idx < len(row):
                max_length = max(max_length, len(str(row[col_idx])))
        adjusted_width = min(max(max_length + COLUMN_WIDTH_PADDING, MIN_COLUMN_WIDTH), MAX_COLUMN_WIDTH)
        worksheet.column_dimensions[column_letter].width = adjusted_width

    return start_row + len(table_data) + TABLE_BOTTOM_SPACING
